#!/usr/bin/env python3
"""
parser_precios.py  v3
Motor inteligente de detección y normalización de listas de precios de proveedores.
Entrada : ruta a archivo .xlsx o .xls
Salida  : JSON stdout  { items:[{codigo,precio,descripcion,hoja}], errores:[], hojas:[] }

Estrategia de detección (por orden):
  1. Header completo: si fila encabezado da TANTO codigo COMO precio → usar directamente
  2. Header parcial : si fila encabezado da codigo pero NO precio → retener codigo del header,
                      buscar precio + descripcion vía heurística sobre columnas numéricas
  3. Heurística pura: si no hay header detectable → inferir todo por tipos de dato
"""

import sys, os, re, json, subprocess, tempfile, shutil, unicodedata, zipfile
import html as _html_mod

# ─── Utilidades ───────────────────────────────────────────────────────────────

def _clean_desc(s):
    if not s: return ''
    return _html_mod.unescape(str(s)).strip()

def norm(v):
    if v is None: return ''
    s = unicodedata.normalize('NFD', str(v))
    s = ''.join(c for c in s if unicodedata.category(c) != 'Mn')
    return re.sub(r'\s+', ' ', s).strip().lower()

def limpiar_precio(v):
    if v is None or v == '': return None
    if isinstance(v, (int, float)):
        return float(v) if 0 < v < 1_000_000_000 else None
    s = re.sub(r'[$\s]', '', str(v).strip())
    if re.search(r'\.\d{3},', s): s = s.replace('.','').replace(',','.')
    elif ',' in s and '.' not in s: s = s.replace(',','.')
    try:
        n = float(s)
        return n if 0 < n < 1_000_000_000 else None
    except: return None

def limpiar_codigo(v):
    if v is None: return None
    s = str(v).strip()
    if re.match(r'^\d+\.0$', s): s = s[:-2]
    s = s.upper()  # normalizar a mayúsculas: la BD guarda los códigos en mayúsculas
    return s if len(s) >= 1 else None

def es_fila_vacia(row):
    return all(c is None or str(c).strip() == '' for c in row)

def clean_meta(v):
    if v is None:
        return ''
    return _html_mod.unescape(str(v)).strip()

FABRICANTES_PERMITIDOS_MYL = {'AP', 'BAIML', 'MYL', 'P01-PORTAFICH'}

def fabricante_myl_permitido(v):
    return clean_meta(v).upper() in FABRICANTES_PERMITIDOS_MYL
# ─── Detección de fila encabezado ─────────────────────────────────────────────

_CLAVES_ENC = ['codigo','cod.','precio','articulo','descripcion',
               'detalle','producto','nombre','importe','referencia']

def es_fila_encabezado(row):
    celdas = [c for c in row if c is not None and str(c).strip()]
    if not celdas: return False
    cortas = sum(1 for c in celdas if len(str(c).strip()) < 25)
    if cortas / len(celdas) < 0.5: return False
    joined = norm(' '.join(str(c) for c in celdas))
    return any(k in joined for k in _CLAVES_ENC)

# ─── Candidatos y matching ────────────────────────────────────────────────────

_CAND_PRECIO = [
    'precio de lista pesos','precio de lista ars','precio de lista',
    'precio lista pesos','precio lista ars','precio lista',
    'precio unitario','precio','importe','valor',
]
_EXCLUIR_PRECIO = ['usd','dolar','dollar','u$s','us$','oferta','minimo','online','costo']

_CAND_CODIGO = [
    'cod_articulo','codigo unico','codigo articulo','codigo',
    'cod. izq.','cod. izq','cod. der','cod.','cod',
    'item','referencia','ref',
]
_CAND_DESC = [
    'descripcion','detalle de los productos','detalle','articulo',
    'nombre','producto','denominacion',
]

def _match(hn, candidates, min_sw=7, min_inc=8):
    if len(hn) > 30: return False
    for c in candidates:
        cn = norm(c)
        if hn == cn: return True
        if len(cn) >= min_sw and hn.startswith(cn): return True
        if len(cn) >= min_inc and cn in hn: return True
    return False

def _precio_excluido(hn):
    return any(e in hn for e in _EXCLUIR_PRECIO)

# ─── Detección de columnas desde header ───────────────────────────────────────

def detectar_columnas_header(header_row):
    """
    Detecta código, precio y descripción desde una fila de encabezado.
    Retorna dict {codigo, precio, descripcion, codigo_alt}.
    Un campo None indica que no se encontró.
    """
    result = {'codigo': None, 'precio': None, 'descripcion': None, 'codigo_alt': None}

    # Precio: recorrer candidatos en orden de especificidad (más específico primero)
    for cand in _CAND_PRECIO:
        cn = norm(cand)
        for idx, cell in enumerate(header_row):
            if cell is None: continue
            hn = norm(str(cell))
            if _precio_excluido(hn): continue
            if hn == cn or (len(cn) >= 7 and hn.startswith(cn)) or (len(cn) >= 8 and cn in hn):
                result['precio'] = idx
                break
        if result['precio'] is not None: break

    # Código y descripción (en un solo recorrido)
    for idx, cell in enumerate(header_row):
        if cell is None: continue
        hn = norm(str(cell))
        if not hn: continue

        if result['codigo'] is None and _match(hn, _CAND_CODIGO):
            result['codigo'] = idx
        elif result['codigo_alt'] is None and _match(hn, _CAND_CODIGO):
            result['codigo_alt'] = idx

        # Descripción: solo si la columna no es ya el código
        if (result['descripcion'] is None
                and _match(hn, _CAND_DESC)
                and idx != result['codigo']
                and idx != result['codigo_alt']):
            result['descripcion'] = idx

    return result

# ─── Detección heurística ─────────────────────────────────────────────────────

def detectar_precio_heuristica(rows, excluir_cols=None):
    """
    Encuentra la columna de precio y descripción por tipos de dato.
    excluir_cols: set de índices a ignorar (ya asignados desde header).
    Retorna dict {precio, descripcion}.
    """
    excluir = excluir_cols or set()
    sample = [r for r in rows if not es_fila_vacia(r)][:40]
    if not sample: return {'precio': None, 'descripcion': None}

    ncols = max((len(r) for r in sample), default=0)
    stats = []

    for c in range(ncols):
        if c in excluir: continue
        vals = [r[c] for r in sample if c < len(r) and r[c] is not None and str(r[c]).strip()]
        if not vals:
            stats.append({'col':c,'precio_score':0,'avg_len':0,'total':0,'avg_num':0}); continue
        ok = [p for p in [limpiar_precio(v) for v in vals] if p is not None]
        stats.append({
            'col': c,
            'precio_score': len(ok)/len(vals),
            'avg_len': sum(len(str(v).strip()) for v in vals)/len(vals),
            'total': len(vals),
            'avg_num': sum(ok)/len(ok) if ok else 0,
        })

    precio_cols = [s for s in stats if s['precio_score'] > 0.7 and s['avg_num'] > 500]
    if not precio_cols: return {'precio': None, 'descripcion': None}
    precio_col = max(precio_cols, key=lambda s: s['avg_num'])

    antes = [s for s in stats
             if s['col'] < precio_col['col']
             and s['total'] >= len(sample) * 0.3
             and s['avg_len'] > 2]

    desc_col = max(antes, key=lambda s: s['avg_len']) if antes else None

    return {
        'precio': precio_col['col'],
        'descripcion': desc_col['col'] if desc_col else None,
    }

def detectar_columnas_heuristica(rows):
    """
    Detección completa cuando no hay header. Infiere codigo, precio, descripcion.
    """
    sample = [r for r in rows if not es_fila_vacia(r)][:40]
    if not sample: return None

    ncols = max((len(r) for r in sample), default=0)
    stats = []
    for c in range(ncols):
        vals = [r[c] for r in sample if c < len(r) and r[c] is not None and str(r[c]).strip()]
        if not vals:
            stats.append({'col':c,'precio_score':0,'avg_len':0,'total':0,'avg_num':0}); continue
        ok = [p for p in [limpiar_precio(v) for v in vals] if p is not None]
        stats.append({
            'col': c,
            'precio_score': len(ok)/len(vals),
            'avg_len': sum(len(str(v).strip()) for v in vals)/len(vals),
            'total': len(vals),
            'avg_num': sum(ok)/len(ok) if ok else 0,
        })

    precio_cols = [s for s in stats if s['precio_score'] > 0.7 and s['avg_num'] > 500]
    if not precio_cols: return None
    precio_col = max(precio_cols, key=lambda s: s['avg_num'])

    antes = [s for s in stats
             if s['col'] < precio_col['col']
             and s['total'] >= len(sample) * 0.3
             and s['avg_len'] > 1]
    if not antes: return None

    desc_col = max(antes, key=lambda s: s['avg_len'])
    codigo_cands = [s for s in antes if s['col'] != desc_col['col']]
    codigo_col = min(codigo_cands, key=lambda s: s['avg_len']) if codigo_cands else desc_col

    return {
        'codigo': codigo_col['col'],
        'precio': precio_col['col'],
        'descripcion': desc_col['col'] if desc_col != codigo_col else None,
        'codigo_alt': None,
    }

# ─── Parsear una hoja ─────────────────────────────────────────────────────────

def parsear_hoja(sheet_name, ws):
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2: return []

    header_idx = -1
    cols = None
    data_rows = [r for r in rows if not es_fila_vacia(r)]

    # Buscar fila header en primeras 12 filas
    for i, row in enumerate(rows[:12]):
        if es_fila_vacia(row): continue
        if not es_fila_encabezado(row): continue
        detected = detectar_columnas_header(row)

        if detected['precio'] is not None and detected['codigo'] is not None:
            # Caso 1: header completo → usar directamente
            header_idx = i
            cols = detected
            break

        if detected['codigo'] is not None and detected['precio'] is None:
            # Caso 2: header parcial (tiene código, sin precio)
            # → retener código del header, buscar precio+desc vía heurística
            excluir = {detected['codigo']}
            if detected['codigo_alt'] is not None: excluir.add(detected['codigo_alt'])
            heur = detectar_precio_heuristica(data_rows, excluir_cols=excluir)
            if heur['precio'] is not None:
                header_idx = i
                cols = {
                    'codigo': detected['codigo'],
                    'codigo_alt': detected['codigo_alt'],
                    'precio': heur['precio'],
                    'descripcion': heur['descripcion'],
                }
                break

    # Caso 3: sin header utilizable → heurística completa
    if cols is None or cols['precio'] is None:
        cols = detectar_columnas_heuristica(data_rows)
        if not cols: return []
        header_idx = -1

    items = []
    for row in rows[header_idx + 1:]:
        if es_fila_vacia(row): continue

        def get(idx):
            return row[idx] if idx is not None and idx < len(row) else None

        precio = limpiar_precio(get(cols['precio']))
        if not precio: continue

        codigos = []
        c1 = limpiar_codigo(get(cols['codigo']))
        if c1: codigos.append(c1)
        c2 = limpiar_codigo(get(cols.get('codigo_alt')))
        if c2 and c2 not in codigos: codigos.append(c2)
        if not codigos: continue

        descripcion = str(get(cols['descripcion']) or '').strip()

        for codigo in codigos:
            items.append({'codigo': codigo, 'precio': precio,
                          'descripcion': _clean_desc(descripcion), 'hoja': sheet_name})
    return items

# ─── Selección de hojas ───────────────────────────────────────────────────────

_HOJAS_IGNORAR = {
    'indice','encabezado','portada',
    'precio minimo online','precios minimos online',
    'hoja2','hoja3','page2','page3',
}

def seleccionar_hojas(sheet_names):
    if len(sheet_names) == 1: return sheet_names
    nn = [(n, norm(n)) for n in sheet_names]

    for name, n in nn:
        if 'lista abreviada' in n: return [name]  # distri

    preferidas = [name for name, n in nn if n.startswith('lista de precio')]
    if preferidas: return preferidas  # LAM, SUDIMAR

    preferidas = [name for name, n in nn if n == 'page1']
    if preferidas: return preferidas  # MYL

    # Hojas que empiezan con 'lista' (ej: 'Lista 166' de Faros Ausili)
    preferidas = [name for name, n in nn if n.startswith('lista')]
    if preferidas: return preferidas

    hoja1 = [name for name, n in nn if n.startswith('hoja1')]
    if hoja1 and len(sheet_names) <= 4: return hoja1  # lider

    return [name for name, n in nn if n not in _HOJAS_IGNORAR]


# ─── Parser específico DISTRIMAR / ALEMARGROUP ───────────────────────────────
# Hoja útil: Lista abreviada
# A = Producto/rubro
# B = Codigo unico    → NO usar
# C = Codigo          → usar para matchear BD
# F = Precio de lista pesos

def _es_archivo_distrimar(ws):
    rows = list(ws.iter_rows(values_only=True, max_row=12))
    for row in rows:
        vals = [norm(c) for c in row if c is not None and str(c).strip()]
        if (
            'codigo unico' in vals
            and 'codigo' in vals
            and 'precio de lista pesos' in vals
        ):
            return True
    return False

def parsear_hoja_distrimar(sheet_name, ws):
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    header_idx = -1
    col_producto = col_codigo = col_precio = None

    for i, row in enumerate(rows[:12]):
        vals = [norm(c) if c is not None else '' for c in row]

        if 'codigo unico' in vals and 'codigo' in vals and 'precio de lista pesos' in vals:
            header_idx = i
            col_producto = vals.index('producto') if 'producto' in vals else None

            # IMPORTANTE: usar columna "Codigo", NO "Codigo unico" ni "Codigo Anterior"
            for idx, val in enumerate(vals):
                if val == 'codigo':
                    col_codigo = idx
                    break

            col_precio = vals.index('precio de lista pesos')
            break

    if header_idx < 0 or col_codigo is None or col_precio is None:
        return []

    items = []

    for row in rows[header_idx + 1:]:
        if not row or es_fila_vacia(row):
            continue

        def get(idx):
            return row[idx] if idx is not None and idx < len(row) else None

        codigo = limpiar_codigo(get(col_codigo))
        precio = limpiar_precio(get(col_precio))

        if not codigo or not precio or precio <= 0:
            continue

        rubro = clean_meta(get(col_producto))

        items.append({
            'codigo':      codigo,
            'precio':      precio,
            'descripcion': _clean_desc(rubro),
            'hoja':        sheet_name,
            'rubro':       _clean_desc(rubro),
            'subrubro':    '',
            'marca':       '',
            'origen':      'DISTRIMAR',
        })

    return items

# ─── Conversión .xls ─────────────────────────────────────────────────────────

def convertir_xls(input_path):
    tmp_dir = tempfile.mkdtemp(prefix='parser_pp_')
    # Ruta absoluta para que funcione cuando lo lanza Node.js (PATH reducido)
    lo_bin = None
    for candidate in ['/usr/bin/libreoffice', '/usr/local/bin/libreoffice',
                      '/opt/libreoffice/program/soffice', '/usr/lib/libreoffice/program/soffice']:
        if os.path.isfile(candidate):
            lo_bin = candidate
            break
    if not lo_bin:
        # Último recurso: buscar en PATH
        import shutil as _shutil
        lo_bin = _shutil.which('libreoffice') or _shutil.which('soffice')
    if not lo_bin:
        shutil.rmtree(tmp_dir, ignore_errors=True)
        raise RuntimeError('No se encontró LibreOffice. Instalá libreoffice con: dnf install libreoffice-calc')

    r = subprocess.run(
        [lo_bin, '--headless', '--convert-to', 'xlsx', '--outdir', tmp_dir, input_path],
        stdout=subprocess.PIPE, stderr=subprocess.PIPE, timeout=120)
    if r.returncode != 0:
        shutil.rmtree(tmp_dir, ignore_errors=True)
        raise RuntimeError('LibreOffice error: ' + (r.stderr or b'').decode('utf-8', errors='replace')[:200])
    out = os.path.join(tmp_dir, os.path.splitext(os.path.basename(input_path))[0] + '.xlsx')
    if not os.path.exists(out):
        shutil.rmtree(tmp_dir, ignore_errors=True)
        raise RuntimeError(f'No se encontró el archivo en {tmp_dir}')
    return out, tmp_dir

# ─── Detección y parser especializado para FAROS AUSILI ──────────────────────
# Su lista tiene:
#   col 0 = código, col 1 = descripción
#   col 2 = precio en PESOS (algunos productos)
#   col 3 = precio en DÓLARES (mayoría de productos) → convertir × TIPO_CAMBIO_USD
# La fila 6 (índice) tiene los encabezados "PESOS" y "DÓLARES" en esas columnas.

TIPO_CAMBIO_USD = float(os.environ.get('TIPO_CAMBIO_USD', '1500'))  # configurable por variable de entorno

def _es_archivo_faros_ausili(ws):
    """True si la hoja tiene la estructura característica de Faros Ausili."""
    rows = list(ws.iter_rows(values_only=True, max_row=10))
    for row in rows:
        if not row: continue
        # Normalizar cada celda quitando acentos para comparar
        vals_norm = set()
        for v in row:
            if v is not None and str(v).strip():
                s = unicodedata.normalize('NFD', str(v).strip())
                s = ''.join(c for c in s if unicodedata.category(c) != 'Mn')
                vals_norm.add(s.upper())
        if 'PESOS' in vals_norm and 'DOLARES' in vals_norm:
            return True
    return False


def parsear_hoja_faros_ausili(sheet_name, ws):
    """
    Parser especializado para FAROS AUSILI.
    - col 0: código
    - col 1: descripción
    - col 2: precio en PESOS
    - col 3: precio en DÓLARES × TIPO_CAMBIO_USD
    - detecta rubro/sección desde filas título
    """
    rows = list(ws.iter_rows(values_only=True))
    items = []
    rubro_actual = ''

    for row in rows:
        if not row or es_fila_vacia(row):
            continue

        def get(idx):
            return row[idx] if idx < len(row) else None

        codigo = limpiar_codigo(get(0))
        descripcion = clean_meta(get(1))

        precio_pesos = limpiar_precio(get(2))
        precio_usd   = limpiar_precio(get(3))

        # Detectar secciones: filas sin precio válido y texto descriptivo
        if not precio_pesos and not precio_usd:
            posible = clean_meta(get(0)) or clean_meta(get(1))
            posible_norm = norm(posible)

            if (
                posible
                and len(posible) >= 4
                and not posible_norm.startswith('codigo')
                and not posible_norm.startswith('cod')
                and any(k in posible_norm for k in ['faro', 'lente', 'optica', 'portal', 'lampara', 'universal'])
            ):
                rubro_actual = posible
            continue

        if not codigo:
            continue

        if precio_pesos and precio_pesos > 0:
            precio_final = precio_pesos
        elif precio_usd and precio_usd > 0:
            precio_final = round(precio_usd * TIPO_CAMBIO_USD, 2)
        else:
            continue

        items.append({
            'codigo':      codigo,
            'precio':      precio_final,
            'descripcion': _clean_desc(descripcion),
            'hoja':        sheet_name,
            'rubro':       _clean_desc(rubro_actual),
            'subrubro':    '',
            'marca':       '',
            'origen':      'AUSILI',
        })

    return items

# ─── Detección y mapeo especial para archivos DM ─────────────────────────────
# DM envía un único archivo con todos sus sub-proveedores mezclados.
# La columna "origen" y "padron" determinan a qué sub-proveedor pertenece cada fila.
# El usuario selecciona "DM" (id=8) y el sistema distribuye automáticamente.

_DM_COLUMNAS = {'cod_articulo', 'padron', 'origen', 'articulo', 'precio', 'marca'}

def _es_archivo_dm(ws):
    """True si la hoja tiene la estructura característica de DM."""
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i > 5: break
        if row and any(c is not None for c in row):
            celdas = {norm(str(c)) for c in row if c is not None and str(c).strip()}
            if len(_DM_COLUMNAS & celdas) >= 4:
                return True
    return False

def _mapear_subproveedor_dm(padron, origen):
    """
    Devuelve lista de proveedor_ids para una fila del archivo DM.
    Basado en los sub-proveedores reales de la BD:
      8  DM            14 DM LAM       18 DM VIDRIOS
      12 DM FAL        15 DM FITAM     19 DM LAMPARAS
      16 DM VIC        17 DM AP        34 DM LAMPARAS LED
    """
    p = padron.strip().upper() if padron else ''
    o = origen.strip().upper() if origen else ''

    provs = []

    # Origen → sub-proveedor específico
    if o == 'VIC':         provs.append(16)   # DM VIC
    if o == 'LAM':         provs.append(14)   # DM LAM
    if o == 'FITAM':       provs.append(15)   # DM FITAM
    if o == 'A-PLASTIC':   provs.append(17)   # DM AP
    if o == 'FAL':         provs.append(12)   # DM FAL
    if o == 'POLICARBONATO' or p == 'VIDRIOS DE OPTICA':
        provs.append(18)                       # DM VIDRIOS
    if p == 'LAMPARAS':
        provs.append(19)                       # DM LAMPARAS
        provs.append(34)                       # DM LAMPARAS LED

    # Todo lo demás va al proveedor DM genérico (id=8)
    if not provs or o in (
        'DEPO','IMPORTADO','NACIONAL','ORIGINAL','T-ORIGINAL','ORIGINAL-A',
        'BRASIL','TAIWAN','ARGENTA','HELLA','INOVOX','MARELLI','VALEO','ORGUS',
        'DAM','CELCO','EURO','HT','LEDEX','COFRAN','BORHAM','PHILIPS',
        'FICO','FTM','METAGAL','VIEW-MAX','SAN JUSTO','','.','BRASIL'
    ):
        provs.append(8)                        # DM genérico

    # Deduplicar manteniendo orden
    seen = set()
    return [p for p in provs if not (p in seen or seen.add(p))]

def parsear_hoja_dm(sheet_name, ws):
    """
    Parser especializado para el archivo DM.
    Devuelve items con campo extra 'proveedor_id_override' (lista de ids).
    Header: cod_articulo | marca | padron | articulo | origen | precio
    """
    rows = list(ws.iter_rows(values_only=True))
    if not rows: return []

    # Buscar fila header
    header_idx = -1
    col_cod = col_padron = col_origen = col_articulo = col_precio = col_marca = None

    for i, row in enumerate(rows[:5]):
        if not row: continue
        cells_norm = [norm(str(c)) if c else '' for c in row]
        if 'cod_articulo' in cells_norm and 'precio' in cells_norm:
            header_idx = i
            col_cod     = cells_norm.index('cod_articulo')
            col_precio  = cells_norm.index('precio')
            col_padron  = cells_norm.index('padron') if 'padron' in cells_norm else None
            col_origen  = cells_norm.index('origen') if 'origen' in cells_norm else None
            col_articulo= cells_norm.index('articulo') if 'articulo' in cells_norm else None
            col_marca   = cells_norm.index('marca') if 'marca' in cells_norm else None
            break

    if header_idx < 0 or col_precio is None: return []

    items = []
    for row in rows[header_idx + 1:]:
        if not row or all(c is None or str(c).strip() == '' for c in row): continue

        def get(idx):
            return row[idx] if idx is not None and idx < len(row) else None

        precio = limpiar_precio(get(col_precio))
        if not precio: continue

        codigo = limpiar_codigo(get(col_cod))
        if not codigo: continue

        padron  = str(get(col_padron)  or '').strip()
        origen  = str(get(col_origen)  or '').strip()
        articulo= str(get(col_articulo) or '').strip()
        marca    = str(get(col_marca) or '').strip()

        prov_ids = _mapear_subproveedor_dm(padron, origen)

        for prov_id in prov_ids:
            items.append({
                'codigo':               codigo,
                'precio':               precio,
                'descripcion':          articulo,
                'hoja':                 sheet_name,
                'rubro':                padron,
                'subrubro':             '',
                'marca':                marca,
                'origen':               origen,
                'proveedor_id_override': prov_id,
            })

    return items


# ─── Detección y parser especializado para FAL (F066) ────────────────────────
# FAL tiene hojas por marca (FORD, RENAULT, etc.) con estructura:
#   col 0 = código largo (ej: '036801') — NO usado, la BD tiene el corto
#   col 1 = código corto (ej: '6801')   — ES el que está en la BD
#   col 2 = descripción
#   col 3 = I/D (indicador, ignorar)
#   col 4 = precio
# La "LISTA PRECIOS GESTION AL DIA" tiene códigos largos — NO coinciden con la BD.

_FAL_HOJAS_MARCA = {
    'ford', 'renault', 'fiat  alfa romeo', 'fiat alfa romeo',
    'peugeot', 'volkswagen', 'chevrolet', 'citroen',
    'nissan jeep bmw suzuki chrysler',
    'honda-toyota-hyundai- audi -mi',
    'honda toyota hyundai audi',
    'm-benz -ivecco -scania', 'mercedes benz iveco scania',
    'universales - rastrojero', 'universales',
}
_FAL_HOJA_MAESTRA = 'lista precios gestion al dia'  # ignorar, usa codigos largos

def _es_hoja_marca_fal(sheet_name_norm):
    """True si la hoja es una hoja de marca de FAL."""
    # Detectar por nombre: cualquier hoja que no sea encabezado/maestra/hoja vacía
    ignorar = {'encabezado', _FAL_HOJA_MAESTRA, 'hoja1', 'hoja2', 'hoja3',
               'page1', 'page2', 'page3'}
    if sheet_name_norm in ignorar: return False
    return True

def _es_archivo_fal(wb_sheetnames):
    """True si el workbook tiene la estructura de FAL (hoja maestra + hojas por marca)."""
    names_norm = {n.strip().lower() for n in wb_sheetnames}
    return _FAL_HOJA_MAESTRA in names_norm


def parsear_hoja_marca_fal(sheet_name, ws):
    """
    Parser para hojas por marca de FAL.
    Mantiene:
    - col 1 = código corto, el que está en BD
    - col 2 = descripción
    - col 4 = precio
    Agrega:
    - rubro detectado por título/sección interna
    """
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    items = []
    rubro_actual = ''

    for row in rows:
        if not row or es_fila_vacia(row):
            continue

        def get(idx):
            return row[idx] if idx < len(row) else None

        codigo_raw = get(1)
        descripcion_raw = clean_meta(get(2))
        precio = limpiar_precio(get(4))

        # Detectar título/sección si no hay precio válido
        if not precio:
            posible = clean_meta(get(0)) or clean_meta(get(1)) or clean_meta(get(2))
            posible_norm = norm(posible)

            if (
                posible
                and len(posible) >= 4
                and not any(ch.isdigit() for ch in posible)
                and any(k in posible_norm for k in ['faro', 'optica', 'lente', 'portal', 'ojo', 'circuito'])
            ):
                rubro_actual = posible
            continue

        if codigo_raw is None:
            continue

        s = str(codigo_raw).strip()
        if re.match(r'^\d+\.0$', s):
            s = s[:-2]

        codigo = s.upper().strip()

        if not codigo or codigo.startswith('*') or codigo in ('CÓDIGO', 'CODIGO', 'COD.'):
            continue

        if not any(c.isdigit() for c in codigo):
            continue

        if not precio or precio <= 0:
            continue

        items.append({
            'codigo':      codigo,
            'precio':      precio,
            'descripcion': _clean_desc(descripcion_raw),
            'hoja':        sheet_name,
            'rubro':       _clean_desc(rubro_actual),
            'subrubro':    '',
            'marca':       sheet_name,
            'origen':      'FAL',
        })
    return items

# ─── Detección y parser especializado para MYL ───────────────────────────────
# MYL envía un archivo con hoja 'Page1' y estructura fija:
#   fila 0: título 'ListaPreciosMYL'
#   fila 1: CODART | NOMMAR | NOMRUB | NOMSUB | DESCRI | PRELIS | OFERTA | FECCAD | IVA
#   col 0 = código de artículo (CODART)
#   col 4 = descripción (DESCRI)
#   col 5 = precio de lista en pesos (PRELIS)

def _es_archivo_myl(ws):
    """True si la hoja tiene la estructura característica de MYL."""
    for i, row in enumerate(ws.iter_rows(values_only=True, max_row=3)):
        if not row: continue
        for cell in row:
            if cell is not None and str(cell).strip().upper() == 'LISTAPRECIOSMYL':
                return True
        vals = [norm(str(c)) for c in row if c is not None and str(c).strip()]
        if 'codart' in vals and 'prelis' in vals:
            return True
    return False


def parsear_hoja_myl(sheet_name, ws):
    """
    Parser especializado para MYL.
    col 0 = CODART (código)
    col 1 = NOMMAR (fabricante) → solo AP, BAIML, MYL y P01-PORTAFICH
    col 2 = NOMRUB (rubro)
    col 3 = NOMSUB (subrubro)
    col 4 = DESCRI (descripción)
    col 5 = PRELIS (precio de lista pesos)
    col 8 = IVA
    """
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    header_idx = -1
    col_cod = col_desc = col_precio = col_marca = col_rubro = col_subrubro = col_iva = None

    for i, row in enumerate(rows[:5]):
        if not row:
            continue
        vals = [norm(str(c)) if c is not None else '' for c in row]
        if 'codart' in vals and 'prelis' in vals:
            header_idx = i
            col_cod     = vals.index('codart')
            col_precio  = vals.index('prelis')
            col_desc    = vals.index('descri') if 'descri' in vals else None
            col_marca   = vals.index('nommar') if 'nommar' in vals else None
            col_rubro   = vals.index('nomrub') if 'nomrub' in vals else None
            col_subrubro= vals.index('nomsub') if 'nomsub' in vals else None
            col_iva     = vals.index('iva') if 'iva' in vals else None
            break

    if header_idx < 0 or col_precio is None:
        return []

    items = []
    for row in rows[header_idx + 1:]:
        if not row or es_fila_vacia(row):
            continue

        def get(idx):
            return row[idx] if idx is not None and idx < len(row) else None

        precio = limpiar_precio(get(col_precio))
        if not precio or precio <= 0:
            continue

        codigo = limpiar_codigo(get(col_cod))
        if not codigo:
            continue

        marca = clean_meta(get(col_marca))
        if not fabricante_myl_permitido(marca):
            continue

        descripcion = clean_meta(get(col_desc))
        rubro = clean_meta(get(col_rubro))
        subrubro = clean_meta(get(col_subrubro))
        iva = limpiar_precio(get(col_iva))

        items.append({
            'codigo':      codigo,
            'precio':      precio,
            'descripcion': _clean_desc(descripcion),
            'hoja':        sheet_name,
            'rubro':       _clean_desc(rubro),
            'subrubro':    _clean_desc(subrubro),
            'marca':       _clean_desc(marca),
            'origen':      'MYL',
            'iva':         iva if iva else None,
        })
    return items

def _parsear_xls_con_xlrd(file_path):
    """
    Parsea el XLS de FAL directamente con xlrd.
    - Lee la hoja maestra 'LISTA PRECIOS GESTION AL DIA' para obtener
      el precio real (col 4) y la descripcion larga (col 3).
    - Lee las hojas de marca para obtener cod_largo (col 0) y cod_corto (col 1).
    - El cod_corto es el que existe en la BD (sin el prefijo de hoja).
    - No deduplica: si el mismo codigo corto aparece en varias hojas de marca,
      emite el precio mas alto (criterio habitual). El bulk en Node actualiza
      todos los registros de la BD con ese codigo, sean izquierdo o derecho.
    """
    import xlrd as _xlrd
    import re as _re

    _HOJA_MAESTRA_FAL = 'lista precios gestion al dia'
    _HOJAS_MARCA_FAL  = {
        'ford', 'renault', 'fiat  alfa romeo', 'peugeot', 'volkswagen',
        'chevrolet', 'citroen ', 'nissan jeep bmw suzuki chrysler',
        'honda-toyota-hyundai- audi -mi ', 'm-benz -ivecco -scania',
        ' universales - rastrojero'
    }

    try:
        wb = _xlrd.open_workbook(file_path)
    except Exception as e:
        return {'items': [], 'hojas': [], 'errores': [f'No se pudo leer el .xls: {e}']}

    # ── 1. Cargar mapa de la hoja maestra: cod_largo → (precio, descripcion) ──
    desc_map   = {}
    precio_map = {}
    for sh_name in wb.sheet_names():
        if norm(sh_name) == _HOJA_MAESTRA_FAL:
            ws_m = wb.sheet_by_name(sh_name)
            for i in range(1, ws_m.nrows):
                cod = str(ws_m.cell_value(i, 0)).strip()
                # Normalizar codigo largo (quitar espacios y .0 final)
                cod = _re.sub(r'\s+', '', cod).upper()
                if cod.endswith('.0'): cod = cod[:-2]
                desc  = str(ws_m.cell_value(i, 3)).strip() if ws_m.ncols > 3 else ''
                precio = ws_m.cell_value(i, 4)              if ws_m.ncols > 4 else 0
                if cod:
                    desc_map[cod]   = desc
                    precio_map[cod] = precio
            break

    # ── 2. Parsear hojas de marca ──────────────────────────────────────────────
    dedup  = {}  # cod_corto_upper → item  (precio mayor gana)
    hojas_usadas = []

    for sh_name in wb.sheet_names():
        if norm(sh_name) not in _HOJAS_MARCA_FAL:
            continue
        hojas_usadas.append(sh_name)
        ws = wb.sheet_by_name(sh_name)

        for i in range(1, ws.nrows):
            row = [ws.cell_value(i, j) for j in range(ws.ncols)]

            # col 0 = codigo largo, col 1 = codigo corto, col 4 = precio (VLOOKUP)
            cod_largo_raw = str(row[0]).strip() if row[0] else ''
            cod_corto_raw = str(row[1]).strip() if len(row) > 1 and row[1] else ''

            # Limpiar codigo corto
            if cod_corto_raw.endswith('.0') and cod_corto_raw[:-2].isdigit():
                cod_corto_raw = cod_corto_raw[:-2]
            cod_corto = _re.sub(r'\s+', '', cod_corto_raw).upper()

            if not cod_corto or not any(c.isdigit() for c in cod_corto):
                continue
            if cod_corto.startswith('*'):
                continue

            # Normalizar codigo largo para el mapa
            cod_largo_key = _re.sub(r'\s+', '', cod_largo_raw).upper()
            if cod_largo_key.endswith('.0'): cod_largo_key = cod_largo_key[:-2]

            # Precio: preferir hoja maestra (VLOOKUP de xlrd no siempre resuelve)
            precio = precio_map.get(cod_largo_key, 0)
            if not precio or precio <= 0:
                precio = limpiar_precio(row[4]) if len(row) > 4 else 0
            if not precio or precio <= 0:
                continue

            # Descripcion: preferir hoja maestra, fallback col 2
            desc = desc_map.get(cod_largo_key, '')
            if not desc and len(row) > 2:
                desc = str(row[2]).strip()

            # Deduplicar por cod_corto, precio mayor gana
            existing = dedup.get(cod_corto)
            if existing is None or precio > existing['precio']:
                dedup[cod_corto] = {
                    'codigo':      cod_corto,
                    'precio':      float(precio),
                    'descripcion': _clean_desc(desc),
                    'hoja':        sh_name,
                    'rubro':       '',
                    'subrubro':    '',
                    'marca':       sh_name,
                    'origen':      'FAL',
                }

    return {
        'items':  list(dedup.values()),
        'hojas':  hojas_usadas,
        'errores': []
    }



# ─── Detección y parser rápido para DISTRIMAR / ALEMARGROUP ──────────────────

def _detectar_distrimar_xlsx(file_path):
    """Detecta DISTRIMAR/Alemar por hoja 'Lista abreviada' y headers clave."""
    try:
        with zipfile.ZipFile(file_path) as zf:
            if 'xl/workbook.xml' not in zf.namelist():
                return False
            wb_xml = zf.read('xl/workbook.xml').decode('utf-8', errors='replace')
            if 'Lista abreviada' not in wb_xml:
                return False
            if 'xl/sharedStrings.xml' not in zf.namelist():
                return False
            chunk = zf.read('xl/sharedStrings.xml')[:250000].decode('utf-8', errors='replace')
            return 'Codigo unico' in chunk and 'Precio de lista pesos' in chunk
    except Exception:
        return False

def _parsear_distrimar_xlsx(file_path):
    """
    Parser rápido de DISTRIMAR/Alemar.
    Usa solo hoja 'Lista abreviada':
    A=Producto/rubro, C=Codigo correcto, F=Precio de lista pesos.
    Ignora B=Codigo unico y E=Precio de lista usd.
    """
    items = []
    _row_re = re.compile(r'<row[^>]*r="(\d+)"[^>]*>(.*?)</row>', re.DOTALL)
    _cell_re_tpl = r'<c r="{col}(\d+)"([^>]*)>(?:<f[^>]*>[^<]*</f>)?(?:<v[^>]*>([^<]*)</v>|<is><t>([^<]*)</t></is>)'
    cell_a = re.compile(_cell_re_tpl.format(col='A'))
    cell_c = re.compile(_cell_re_tpl.format(col='C'))
    cell_f = re.compile(_cell_re_tpl.format(col='F'))

    def cell_value(match, shared):
        if not match:
            return ''
        attrs = match.group(2) or ''
        val = match.group(3) if match.group(3) is not None else (match.group(4) or '')
        if ('t="s"' in attrs or "t='s'" in attrs) and val != '':
            try:
                return shared[int(val)].strip()
            except Exception:
                return ''
        return str(val).strip()

    try:
        with zipfile.ZipFile(file_path) as zf:
            shared = []
            if 'xl/sharedStrings.xml' in zf.namelist():
                ss = zf.read('xl/sharedStrings.xml').decode('utf-8', errors='replace')
                for block in ss.split('<si>')[1:]:
                    texts = re.findall(r'<t[^>]*>([^<]*)</t>', block)
                    shared.append(_html_mod.unescape(''.join(texts)).strip())

            wb_xml = zf.read('xl/workbook.xml').decode('utf-8', errors='replace')
            sheet_entries = re.findall(r'<sheet[^>]+name="([^"]+)"[^>]+r:id="([^"]+)"', wb_xml)
            rels_xml = zf.read('xl/_rels/workbook.xml.rels').decode('utf-8', errors='replace')
            rels = {}
            for rel_m in re.finditer(r'<Relationship\b([^>]+)>', rels_xml):
                attrs = rel_m.group(1)
                id_m = re.search(r'\bId="([^"]+)"', attrs)
                tgt_m = re.search(r'\bTarget="([^"]+)"', attrs)
                if id_m and tgt_m:
                    rels[id_m.group(1)] = tgt_m.group(1)

            sheet_path = None
            for name, rid in sheet_entries:
                if norm(name) == 'lista abreviada':
                    target = rels.get(rid, '').lstrip('/')
                    sheet_path = target if target.startswith('xl/') else f'xl/{target}'
                    break
            if not sheet_path or sheet_path not in zf.namelist():
                return items

            sheet_xml = zf.read(sheet_path).decode('utf-8', errors='replace')

        for m in _row_re.finditer(sheet_xml):
            row_num = int(m.group(1))
            if row_num <= 4:
                continue
            row_xml = m.group(2)

            rubro = cell_value(cell_a.search(row_xml), shared)
            codigo = limpiar_codigo(cell_value(cell_c.search(row_xml), shared))
            precio = limpiar_precio(cell_value(cell_f.search(row_xml), shared))

            if not codigo or not precio or precio <= 0:
                continue

            items.append({
                'codigo':      codigo,
                'precio':      precio,
                'descripcion': _clean_desc(rubro),
                'hoja':        'Lista abreviada',
                'rubro':       _clean_desc(rubro),
                'subrubro':    '',
                'marca':       '',
                'origen':      'DISTRIMAR',
            })
    except Exception:
        pass

    return items

# ─── Detección y parser especializado para CROMOSOL ──────────────────────────
# Su lista tiene 60k+ filas con columnas que incluyen strings enormes (col J).
# Leer con openpyxl tarda 5+ segundos. Este parser usa zipfile + regex sobre
# el XML interno del xlsx, leyendo SOLO columnas B (Código) y D (Precio Neto).
# Resultado: de 5.7s → 0.9s para 63.000 filas.

def _detectar_cromosol(file_path):
    """
    Detecta el formato CROMOSOL leyendo solo sharedStrings.xml con regex.
    No abre openpyxl — mucho más rápido.
    """
    try:
        with zipfile.ZipFile(file_path) as zf:
            if 'xl/sharedStrings.xml' not in zf.namelist():
                return False
            with zf.open('xl/sharedStrings.xml') as f:
                # Leer solo los primeros 2KB para detectar el header
                chunk = f.read(2048).decode('utf-8', errors='replace')
            # El header de CROMOSOL tiene estas columnas exactas
            return ('Empresa' in chunk and 'Precio Neto' in chunk
                    and ('C\u00f3digo' in chunk or 'Código' in chunk or 'Codigo' in chunk))
    except Exception:
        return False


def _parsear_cromosol_xlsx(file_path):
    """
    Lee columnas B (Código), C (Descripción), D (Precio Neto) y G (Rubro)
    usando zipfile + regex. Evita cargar todo el Excel con openpyxl.
    """
    items = []
    _cell_b = re.compile(r'<c r="B\d+"([^>]*)><v>([^<]+)</v>')
    _cell_c = re.compile(r'<c r="C\d+"([^>]*)><v[^>]*>([^<]+)</v>')
    _cell_d = re.compile(r'<c r="D\d+"([^>]*)><v>([^<]+)</v>')
    _cell_g = re.compile(r'<c r="G\d+"([^>]*)><v[^>]*>([^<]+)</v>')
    _row_re = re.compile(r'<row[^>]*r="(\d+)"[^>]*>(.*?)</row>', re.DOTALL)

    def read_cell(match, shared):
        if not match:
            return ''
        attrs, val = match.group(1), match.group(2)
        if 't="s"' in attrs or "t='s'" in attrs:
            try:
                return shared[int(val)].strip()
            except Exception:
                return ''
        return val.strip()

    try:
        with zipfile.ZipFile(file_path) as zf:
            shared = []
            if 'xl/sharedStrings.xml' in zf.namelist():
                with zf.open('xl/sharedStrings.xml') as f:
                    ss = f.read().decode('utf-8', errors='replace')
                for block in ss.split('<si>')[1:]:
                    texts = re.findall(r'<t[^>]*>([^<]*)</t>', block)
                    shared.append(''.join(texts))

            sheets = [n for n in zf.namelist()
                      if n.startswith('xl/worksheets/sheet') and n.endswith('.xml')]
            if not sheets:
                return items
            with zf.open(sheets[0]) as f:
                sheet = f.read().decode('utf-8', errors='replace')

        for m in _row_re.finditer(sheet):
            if int(m.group(1)) == 1:
                continue
            row_xml = m.group(2)

            mb = _cell_b.search(row_xml)
            md = _cell_d.search(row_xml)
            if not mb or not md:
                continue

            codigo = read_cell(mb, shared)
            if re.match(r'^\d+\.0$', codigo):
                codigo = codigo[:-2]

            precio = limpiar_precio(read_cell(md, shared))
            if not precio or precio <= 0:
                continue

            codigo_clean = limpiar_codigo(codigo)
            if not codigo_clean:
                continue

            desc = read_cell(_cell_c.search(row_xml), shared)
            rubro = read_cell(_cell_g.search(row_xml), shared)

            items.append({
                'codigo':      codigo_clean,
                'precio':      precio,
                'descripcion': _clean_desc(desc),
                'hoja':        'lista precios',
                'rubro':       _clean_desc(rubro),
                'subrubro':    '',
                'marca':       '',
                'origen':      'CROMOSOL',
            })
    except Exception:
        pass

    return items

# ─── Detección y parser especializado para MYL ───────────────────────────────
# MYL tiene 34k+ filas. Col A=CODART (código), Col F=PRELIS (precio numérico).
# Fila 1 = título 'ListaPreciosMYL', fila 2 = header, datos desde fila 3.
# Los precios son numéricos directos (no shared strings) → aún más rápido.

def _detectar_myl(file_path):
    """Detecta MYL leyendo los primeros 1KB de sharedStrings."""
    try:
        with zipfile.ZipFile(file_path) as zf:
            if 'xl/sharedStrings.xml' not in zf.namelist():
                return False
            with zf.open('xl/sharedStrings.xml') as f:
                chunk = f.read(1024).decode('utf-8', errors='replace')
            return 'CODART' in chunk and 'PRELIS' in chunk
    except Exception:
        return False


def _parsear_myl_xlsx(file_path):
    """
    Lee MYL por XML:
    A=CODART, B=NOMMAR, C=NOMRUB, D=NOMSUB, E=DESCRI, F=PRELIS, I=IVA.
    Solo procesa fabricantes permitidos: AP, BAIML, MYL y P01-PORTAFICH.
    """
    _cell_a = re.compile(r'<c r="A(\d+)"([^>]*)><v>([^<]+)</v>')
    _cell_b = re.compile(r'<c r="B(\d+)"([^>]*)><v[^>]*>([^<]+)</v>')
    _cell_c = re.compile(r'<c r="C(\d+)"([^>]*)><v[^>]*>([^<]+)</v>')
    _cell_d = re.compile(r'<c r="D(\d+)"([^>]*)><v[^>]*>([^<]+)</v>')
    _cell_e = re.compile(r'<c r="E(\d+)"([^>]*)><v[^>]*>([^<]+)</v>')
    _cell_f = re.compile(r'<c r="F(\d+)"([^>]*)><v>([^<]+)</v>')
    _cell_i = re.compile(r'<c r="I(\d+)"([^>]*)><v[^>]*>([^<]+)</v>')
    _row_re = re.compile(r'<row[^>]*r="(\d+)"[^>]*>(.*?)</row>', re.DOTALL)
    items = []

    try:
        with zipfile.ZipFile(file_path) as zf:
            shared = []
            if 'xl/sharedStrings.xml' in zf.namelist():
                with zf.open('xl/sharedStrings.xml') as f:
                    ss = f.read().decode('utf-8', errors='replace')
                for block in ss.split('<si>')[1:]:
                    texts = re.findall(r'<t[^>]*>([^<]*)</t>', block)
                    shared.append(''.join(texts))

            sheets = [n for n in zf.namelist()
                      if n.startswith('xl/worksheets/sheet') and n.endswith('.xml')]
            if not sheets:
                return items
            with zf.open(sheets[0]) as f:
                sheet = f.read().decode('utf-8', errors='replace')

        def read_cell_text(regex_obj, row_xml):
            mcell = regex_obj.search(row_xml)
            if not mcell:
                return ''
            attrs, val = mcell.group(2), mcell.group(3)
            if 't="s"' in attrs or "t='s'" in attrs:
                try:
                    return shared[int(val)].strip()
                except Exception:
                    return ''
            return val.strip()

        for m in _row_re.finditer(sheet):
            if int(m.group(1)) <= 2:
                continue
            row_xml = m.group(2)

            ma = _cell_a.search(row_xml)
            mf = _cell_f.search(row_xml)
            if not ma or not mf:
                continue

            attrs_a, val_a = ma.group(2), ma.group(3)
            if 't="s"' in attrs_a or "t='s'" in attrs_a:
                try:
                    codigo = shared[int(val_a)].strip()
                except Exception:
                    continue
            else:
                codigo = val_a.strip()
                if re.match(r'^\d+\.0$', codigo):
                    codigo = codigo[:-2]

            try:
                precio = float(mf.group(3))
            except Exception:
                continue

            codigo_clean = limpiar_codigo(codigo)
            if not codigo_clean or precio <= 0:
                continue

            fabricante = read_cell_text(_cell_b, row_xml)
            if not fabricante_myl_permitido(fabricante):
                continue

            rubro = read_cell_text(_cell_c, row_xml)
            subrubro = read_cell_text(_cell_d, row_xml)
            desc = read_cell_text(_cell_e, row_xml)
            iva = limpiar_precio(read_cell_text(_cell_i, row_xml))

            items.append({
                'codigo':      codigo_clean,
                'precio':      precio,
                'descripcion': _clean_desc(desc),
                'hoja':        'Page1',
                'rubro':       _clean_desc(rubro),
                'subrubro':    _clean_desc(subrubro),
                'marca':       _clean_desc(fabricante),
                'origen':      'MYL',
                'iva':         iva if iva else None,
            })
    except Exception:
        pass

    return items

# ─── Detección y parser especializado para LIDERCAR ──────────────────────────
# LIDERCAR tiene ~10k filas. Header en fila 6, datos desde fila 7.
# Col A=Cód.Izq, Col B=Cód.der (ambos pueden tener código), Col I=Precio lista.
# Los códigos son numéricos en el XML (no shared strings) → carga mínima.

def _detectar_lidercar(file_path):
    """Detecta LIDERCAR buscando sus encabezados en sharedStrings."""
    try:
        with zipfile.ZipFile(file_path) as zf:
            if 'xl/sharedStrings.xml' not in zf.namelist():
                return False
            with zf.open('xl/sharedStrings.xml') as f:
                ss = f.read().decode('utf-8', errors='replace')
            return ('Cód. Izq' in ss or 'C\u00f3d. Izq' in ss) and 'Precio lista' in ss
    except Exception:
        return False


def _parsear_lidercar_xlsx(file_path):
    """
    Lee col A (Cód.Izq), col B (Cód.der), col C (Rubro), col D (Descripción)
    y col I (Precio lista) con zipfile+regex.
    Ambas columnas de código pueden tener valor para el mismo producto.
    """
    _cell_a = re.compile(r'<c r="A(\d+)"([^>]*)><v>([^<]+)</v>')
    _cell_b = re.compile(r'<c r="B(\d+)"([^>]*)><v>([^<]+)</v>')
    _cell_c = re.compile(r'<c r="C(\d+)"([^>]*)><v[^>]*>([^<]+)</v>')
    _cell_d = re.compile(r'<c r="D(\d+)"([^>]*)><v[^>]*>([^<]+)</v>')
    _cell_i = re.compile(r'<c r="I(\d+)"([^>]*)><v>([^<]+)</v>')
    _row_re = re.compile(r'<row[^>]*r="(\d+)"[^>]*>(.*?)</row>', re.DOTALL)
    items = []

    try:
        with zipfile.ZipFile(file_path) as zf:
            shared = []
            if 'xl/sharedStrings.xml' in zf.namelist():
                with zf.open('xl/sharedStrings.xml') as f:
                    ss = f.read().decode('utf-8', errors='replace')
                for block in ss.split('<si>')[1:]:
                    texts = re.findall(r'<t[^>]*>([^<]*)</t>', block)
                    shared.append(''.join(texts).strip())

            sheets = [n for n in zf.namelist()
                      if n.startswith('xl/worksheets/sheet') and n.endswith('.xml')]
            if not sheets:
                return items
            with zf.open(sheets[0]) as f:
                sheet = f.read().decode('utf-8', errors='replace')

        def read_cell(match):
            if not match:
                return ''
            attrs, val = match.group(2), match.group(3)
            if 't="s"' in attrs or "t='s'" in attrs:
                try:
                    return shared[int(val)].strip()
                except Exception:
                    return ''
            return val.strip()

        for m in _row_re.finditer(sheet):
            if int(m.group(1)) <= 6:
                continue
            row_xml = m.group(2)

            mi = _cell_i.search(row_xml)
            if not mi:
                continue
            try:
                precio = float(mi.group(3))
            except Exception:
                continue
            if precio <= 0:
                continue

            codigos = []
            for mc in [_cell_a.search(row_xml), _cell_b.search(row_xml)]:
                cod = read_cell(mc)
                if re.match(r'^\d+\.0$', cod):
                    cod = cod[:-2]
                if cod:
                    codigos.append(cod)

            desc = read_cell(_cell_d.search(row_xml))
            rubro = read_cell(_cell_c.search(row_xml))

            for cod in codigos:
                codigo_clean = limpiar_codigo(cod)
                if codigo_clean:
                    items.append({
                        'codigo':      codigo_clean,
                        'precio':      precio,
                        'descripcion': _clean_desc(desc),
                        'hoja':        'Hoja1',
                        'rubro':       _clean_desc(rubro),
                        'subrubro':    '',
                        'marca':       '',
                        'origen':      'LIDERCAR',
                    })
    except Exception:
        pass

    return items

# ─── Detección y parser especializado para FAL (xlsx convertido por Node) ─────
# El .xls de FAL (5MB) es convertido a .xlsx por Node antes de llegar al parser.
# El xlsx resultante tiene strings INLINE (no shared strings) y los precios
# son resultados de fórmulas VLOOKUP pre-calculados.
# Estructura: col B = código corto (el que está en BD), col E = precio.
# El parser lee solo las hojas de marca, ignorando la hoja maestra y ENCABEZADO.

_FAL_HOJAS_MARCA_SET = {
    'ford', 'renault', 'fiat  alfa romeo', 'peugeot', 'volkswagen',
    'chevrolet', 'citroen ', 'nissan jeep bmw suzuki chrysler',
    'honda-toyota-hyundai- audi -mi ', 'm-benz -ivecco -scania',
    ' universales - rastrojero'
}

def _detectar_fal_xlsx(file_path):
    """
    Detecta el xlsx de FAL (convertido por Node o por LibreOffice).
    Criterio principal: tiene la hoja maestra 'LISTA PRECIOS GESTION AL DIA'
    y al menos 3 hojas de marca conocidas.
    Funciona tanto si tiene sharedStrings (Node/xlsx) como si no (LibreOffice).
    """
    try:
        with zipfile.ZipFile(file_path) as zf:
            if 'xl/workbook.xml' not in zf.namelist():
                return False
            with zf.open('xl/workbook.xml') as f:
                wb_xml = f.read().decode('utf-8', errors='replace')
            sheet_names_lower = {n.lower() for n in re.findall(r'<sheet[^>]+name="([^"]+)"', wb_xml)}
            # Debe tener la hoja maestra característica de FAL
            if _FAL_HOJA_MAESTRA not in sheet_names_lower:
                return False
            # Y al menos 3 hojas de marca conocidas
            matches = sheet_names_lower & _FAL_HOJAS_MARCA_SET
            return len(matches) >= 3
    except Exception:
        return False

def _parsear_fal_xlsx(file_path):
    """
    Lee el xlsx de FAL (convertido por Node o LibreOffice).
    - Primero carga la hoja maestra 'LISTA PRECIOS GESTION AL DIA' para obtener
      descripcion larga (col D) y precio (col E) indexados por codigo largo (col A).
    - Luego lee cada hoja de marca:
        col A = codigo largo (para join con hoja maestra)
        col B = codigo corto (el que está en BD)
        col E = precio (resultado del VLOOKUP pre-calculado)
    Soporta tanto strings inline (t="str", sin sharedStrings) como
    strings compartidos (t="s", con sharedStrings.xml).
    """
    _cell_a  = re.compile(r'<c r="A\d+"[^>]*>(?:<f>[^<]*</f>)?(?:<v[^>]*>([^<]*)</v>|<is><t>([^<]*)</t></is>)')
    _cell_b  = re.compile(r'<c r="B\d+"[^>]*>(?:<f>[^<]*</f>)?(?:<v[^>]*>([^<]+)</v>|<is><t>([^<]+)</t></is>)')
    _cell_b_full = re.compile(r'<c r="B\d+"([^>]*)>(?:<f>[^<]*</f>)?(?:<v[^>]*>([^<]+)</v>|<is><t>([^<]+)</t></is>)')
    _cell_c  = re.compile(r'<c r="C\d+"[^>]*>(?:<f>[^<]*</f>)?(?:<v[^>]*>([^<]*)</v>|<is><t>([^<]*)</t></is>)')
    _cell_c_full = re.compile(r'<c r="C\d+"([^>]*)>(?:<f>[^<]*</f>)?(?:<v[^>]*>([^<]*)</v>|<is><t>([^<]*)</t></is>)')
    _cell_d  = re.compile(r'<c r="D\d+"[^>]*>(?:<f>[^<]*</f>)?(?:<v[^>]*>([^<]*)</v>|<is><t>([^<]*)</t></is>)')
    _cell_e  = re.compile(r'<c r="E\d+"[^>]*>(?:<f>[^<]*</f>)?(?:<v[^>]*>([^<]+)</v>|<is><t>([^<]+)</t></is>)')
    _cell_e_full = re.compile(r'<c r="E\d+"([^>]*)>(?:<f>[^<]*</f>)?(?:<v[^>]*>([^<]+)</v>|<is><t>([^<]+)</t></is>)')
    _row_re  = re.compile(r'<row[^>]*r="(\d+)"[^>]*>(.*?)</row>', re.DOTALL)
    _HOJA_MAESTRA = 'lista precios gestion al dia'
    dedup = {}

    def _cell_val(match, shared):
        """Extrae valor de un match: soporta <v>, inlineStr y shared strings."""
        if match is None:
            return ''
        groups = match.groups()
        # Para matches con attrs (full): groups = (attrs, v_val o None, is_val o None)
        # Para matches sin attrs:        groups = (v_val o None, is_val o None)
        if len(groups) == 3:
            attrs, v_val, is_val = groups
        else:
            attrs, v_val, is_val = None, groups[0], groups[1] if len(groups) > 1 else None

        val = v_val if v_val is not None else (is_val or '')
        if attrs and ('t="s"' in attrs or "t='s'" in attrs) and shared:
            try:
                return shared[int(val)]
            except Exception:
                return val
        return val

    try:
        with zipfile.ZipFile(file_path) as zf:
            # Cargar shared strings si existen (Node-converted xlsx los tiene)
            shared = []
            if 'xl/sharedStrings.xml' in zf.namelist():
                with zf.open('xl/sharedStrings.xml') as f:
                    ss = f.read().decode('utf-8', errors='replace')
                for block in ss.split('<si>')[1:]:
                    texts = re.findall(r'<t[^>]*>([^<]*)</t>', block)
                    shared.append(''.join(texts))

            with zf.open('xl/workbook.xml') as f:
                wb_xml = f.read().decode('utf-8', errors='replace')
            sheet_entries = re.findall(r'<sheet[^>]+name="([^"]+)"[^>]+r:id="([^"]+)"', wb_xml)

            with zf.open('xl/_rels/workbook.xml.rels') as f:
                rels_xml = f.read().decode('utf-8', errors='replace')
            # Parsear rels soportando cualquier orden de atributos
            rels = {}
            for rel_m in re.finditer(r'<Relationship\b([^>]+)>', rels_xml):
                attrs = rel_m.group(1)
                id_m  = re.search(r'\bId="([^"]+)"', attrs)
                tgt_m = re.search(r'\bTarget="([^"]+)"', attrs)
                if id_m and tgt_m:
                    rels[id_m.group(1)] = tgt_m.group(1)

            def _resolve_path(target):
                """Convierte target del rels a path dentro del zip."""
                # Rutas absolutas: /xl/worksheets/sheet1.xml → xl/worksheets/sheet1.xml
                t = target.lstrip('/')
                if not t.startswith('xl/'):
                    t = f"xl/{t}"
                return t

            def _get_sheet_xml(name_lower):
                for sh_name, rid in sheet_entries:
                    if sh_name.lower() == name_lower:
                        target = rels.get(rid, '')
                        sheet_path = _resolve_path(target)
                        if sheet_path in zf.namelist():
                            with zf.open(sheet_path) as f:
                                return f.read().decode('utf-8', errors='replace')
                return None

            # ── 1. Cargar hoja maestra: cod_largo → (descripcion, precio) ──────
            desc_map  = {}
            precio_map = {}
            maestra_xml = _get_sheet_xml(_HOJA_MAESTRA)
            if maestra_xml:
                _cell_a_m  = re.compile(r'<c r="A\d+"([^>]*)>(?:<f>[^<]*</f>)?(?:<v[^>]*>([^<]*)</v>|<is><t>([^<]*)</t></is>)')
                _cell_d_m  = re.compile(r'<c r="D\d+"([^>]*)>(?:<f>[^<]*</f>)?(?:<v[^>]*>([^<]*)</v>|<is><t>([^<]*)</t></is>)')
                _cell_e_m  = re.compile(r'<c r="E\d+"([^>]*)>(?:<f>[^<]*</f>)?(?:<v[^>]*>([^<]+)</v>|<is><t>([^<]+)</t></is>)')
                for m in _row_re.finditer(maestra_xml):
                    if int(m.group(1)) <= 1: continue
                    row_xml = m.group(2)
                    ma = _cell_a_m.search(row_xml)
                    md = _cell_d_m.search(row_xml)
                    me = _cell_e_m.search(row_xml)
                    if ma:
                        cod_largo = _cell_val(ma, shared)
                        cod_largo = re.sub(r'\s+', '', str(cod_largo)).upper()
                        if cod_largo.endswith('.0'): cod_largo = cod_largo[:-2]
                        if cod_largo:
                            if md:
                                desc_map[cod_largo] = str(_cell_val(md, shared)).strip()
                            if me:
                                try:
                                    precio_val = float(str(_cell_val(me, shared)).strip())
                                    if precio_val > 0:
                                        precio_map[cod_largo] = precio_val
                                except Exception:
                                    pass

            # ── 2. Parsear hojas de marca ──────────────────────────────────────
            for sheet_name, rid in sheet_entries:
                if sheet_name.lower() not in _FAL_HOJAS_MARCA_SET:
                    continue
                target = rels.get(rid, '')
                sheet_path = _resolve_path(target)
                if sheet_path not in zf.namelist():
                    continue

                with zf.open(sheet_path) as f:
                    sheet_xml = f.read().decode('utf-8', errors='replace')

                _cell_b_re = re.compile(r'<c r="B\d+"([^>]*)>(?:<f>[^<]*</f>)?(?:<v[^>]*>([^<]+)</v>|<is><t>([^<]+)</t></is>)')
                _cell_a_re = re.compile(r'<c r="A\d+"([^>]*)>(?:<f>[^<]*</f>)?(?:<v[^>]*>([^<]*)</v>|<is><t>([^<]*)</t></is>)')
                _cell_c_re = re.compile(r'<c r="C\d+"([^>]*)>(?:<f>[^<]*</f>)?(?:<v[^>]*>([^<]*)</v>|<is><t>([^<]*)</t></is>)')
                _cell_e_re = re.compile(r'<c r="E\d+"([^>]*)>(?:<f>[^<]*</f>)?(?:<v[^>]*>([^<]+)</v>|<is><t>([^<]+)</t></is>)')

                for m in _row_re.finditer(sheet_xml):
                    if int(m.group(1)) <= 1:
                        continue
                    row_xml = m.group(2)

                    # col B = codigo corto
                    mb = _cell_b_re.search(row_xml)
                    if not mb:
                        continue
                    cod_raw = re.sub(r'\s+', '', str(_cell_val(mb, shared)))
                    if re.match(r'^\d+\.0$', cod_raw):
                        cod_raw = cod_raw[:-2]
                    if not cod_raw or cod_raw.startswith('*'):
                        continue
                    if not any(c.isdigit() for c in cod_raw):
                        continue

                    # col A = codigo largo (para join con hoja maestra)
                    ma = _cell_a_re.search(row_xml)
                    cod_largo_key = ''
                    if ma:
                        cod_largo_key = re.sub(r'\s+', '', str(_cell_val(ma, shared))).upper()
                        if cod_largo_key.endswith('.0'): cod_largo_key = cod_largo_key[:-2]

                    # Precio: preferir hoja maestra, fallback col E
                    precio = precio_map.get(cod_largo_key, 0)
                    if not precio or precio <= 0:
                        me = _cell_e_re.search(row_xml)
                        if me:
                            try:
                                precio = float(str(_cell_val(me, shared)).strip())
                            except Exception:
                                precio = 0
                    if not precio or precio <= 0:
                        continue

                    codigo_clean = limpiar_codigo(cod_raw)
                    if not codigo_clean:
                        continue

                    # Descripcion: hoja maestra via cod_largo, fallback col C
                    desc = desc_map.get(cod_largo_key, '')
                    if not desc:
                        mc = _cell_c_re.search(row_xml)
                        if mc:
                            desc = str(_cell_val(mc, shared)).strip()

                    # Deduplicar por codigo_clean, precio mayor gana
                    existing = dedup.get(codigo_clean)
                    if existing is None or precio > existing['precio']:
                        dedup[codigo_clean] = {
                            'codigo':      codigo_clean,
                            'precio':      float(precio),
                            'descripcion': _clean_desc(desc),
                            'hoja':        sheet_name,
                            'rubro':       '',
                            'subrubro':    '',
                            'marca':       sheet_name,
                            'origen':      'FAL',
                        }
    except Exception:
        pass

    return list(dedup.values())


def parsear_archivo(file_path):
    """
    Parsea un archivo Excel de lista de precios.
    Retorna: { items, hojas, errores }
    Cada item: { codigo:str, precio:float, descripcion:str, hoja:str }
    """
    errores = []
    tmp_dir = None
    path_usar = file_path

    # Detectar formato por magic bytes, no por extensión.
    # OLE2 (.xls): D0 CF 11 E0  |  ZIP/xlsx: 50 4B 03 04
    # Así funciona aunque el archivo llegue sin extensión o con nombre raro.
    def _es_xls_ole2(path):
        try:
            with open(path, 'rb') as fh:
                return fh.read(4) == b'\xd0\xcf\x11\xe0'
        except Exception:
            return False

    ext = os.path.splitext(file_path)[1].lower()
    necesita_conversion = (ext == '.xls') or _es_xls_ole2(file_path)

    if necesita_conversion:
        # Intentar convertir con LibreOffice
        try:
            path_usar, tmp_dir = convertir_xls(file_path)
        except Exception as lo_err:
            # LibreOffice no disponible — intentar con xlrd directamente
            try:
                import xlrd as _xlrd
                return _parsear_xls_con_xlrd(file_path)
            except ImportError:
                return {'items': [], 'hojas': [], 'errores': [
                    f'No se pudo convertir el .xls: {lo_err}. Instalá xlrd: pip install xlrd'
                ]}
            except Exception as e:
                return {'items': [], 'hojas': [], 'errores': [str(e)]}

    # ── Detección rápida para listas grandes (CROMOSOL, MYL) ───────────────
    # Estos proveedores tienen 30k-60k filas. Usamos zipfile+regex para leer
    # solo las columnas necesarias, sin cargar openpyxl (5x-8x más rápido).
    if not necesita_conversion:
        if _detectar_distrimar_xlsx(path_usar):
            items_raw = _parsear_distrimar_xlsx(path_usar)
            hoja_nombre = 'Lista abreviada'
        elif _detectar_cromosol(path_usar):
            items_raw = _parsear_cromosol_xlsx(path_usar)
            hoja_nombre = 'lista precios'
        elif _detectar_myl(path_usar):
            items_raw = _parsear_myl_xlsx(path_usar)
            hoja_nombre = 'Page1'
        elif _detectar_lidercar(path_usar):
            items_raw = _parsear_lidercar_xlsx(path_usar)
            hoja_nombre = 'Hoja1'
        elif _detectar_fal_xlsx(path_usar):
            items_raw = _parsear_fal_xlsx(path_usar)
            hoja_nombre = 'FAL'
        else:
            items_raw = None
            hoja_nombre = None

        if items_raw is not None:
            if tmp_dir: shutil.rmtree(tmp_dir, ignore_errors=True)
            dedup = {}
            for it in items_raw:
                k = it['codigo'].strip().upper()
                if k not in dedup or it['precio'] > dedup[k]['precio']:
                    dedup[k] = it
            return {'items': list(dedup.values()), 'hojas': [hoja_nombre], 'errores': []}

    try:
        import openpyxl as _openpyxl
        wb = _openpyxl.load_workbook(path_usar, read_only=True, data_only=True)
    except Exception as e:
        if tmp_dir:
            shutil.rmtree(tmp_dir, ignore_errors=True)
        return {'items': [], 'hojas': [], 'errores': [f'No se pudo leer: {e}']}

    # Detectar si es archivo FAL a nivel de workbook (tiene hoja maestra)
    es_fal = _es_archivo_fal(wb.sheetnames)

    if es_fal:
        # FAL: procesar hojas por marca (no la hoja maestra, que tiene códigos largos)
        hojas = [n for n in wb.sheetnames
                 if norm(n.strip()) not in (_FAL_HOJA_MAESTRA, 'encabezado', 'hoja1', 'hoja2', 'hoja3')]
    else:
        hojas = seleccionar_hojas(wb.sheetnames)

    todos = []
    es_dm = False

    for sh in hojas:
        ws = wb[sh]
        try:
            if es_fal:
                ws2 = wb[sh]
                todos.extend(parsear_hoja_marca_fal(sh, ws2))
            elif _es_archivo_dm(ws):
                es_dm = True
                ws2 = wb[sh]
                todos.extend(parsear_hoja_dm(sh, ws2))
            elif _es_archivo_faros_ausili(ws):
                ws2 = wb[sh]
                todos.extend(parsear_hoja_faros_ausili(sh, ws2))
            elif _es_archivo_distrimar(ws):
                ws2 = wb[sh]
                todos.extend(parsear_hoja_distrimar(sh, ws2))
            elif _es_archivo_myl(ws):
                ws2 = wb[sh]
                todos.extend(parsear_hoja_myl(sh, ws2))
            else:
                todos.extend(parsear_hoja(sh, ws))
        except Exception as e:
            errores.append(f'Hoja "{sh}": {e}')
    wb.close()
    if tmp_dir:
        shutil.rmtree(tmp_dir, ignore_errors=True)

    if es_dm:
        # Para DM: deduplicar por (codigo, proveedor_id_override) — precio mayor
        dedup = {}
        for item in todos:
            key = (item['codigo'].strip().upper(), item.get('proveedor_id_override', 0))
            if key not in dedup or item['precio'] > dedup[key]['precio']:
                dedup[key] = item
    else:
        # Normal: deduplicar solo por codigo → mayor precio
        dedup = {}
        for item in todos:
            key = item['codigo'].strip().upper()
            if key not in dedup or item['precio'] > dedup[key]['precio']:
                dedup[key] = item

    return {'items': list(dedup.values()), 'hojas': hojas, 'errores': errores}

# ─── CLI ──────────────────────────────────────────────────────────────────────

if __name__ == '__main__':
    if len(sys.argv) < 2:
        print(json.dumps({'error': 'Uso: parser_precios.py <archivo>'}))
        sys.exit(1)
    print(json.dumps(parsear_archivo(sys.argv[1]), ensure_ascii=False))